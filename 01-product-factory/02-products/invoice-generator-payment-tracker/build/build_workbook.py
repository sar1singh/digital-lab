import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import date

wb = openpyxl.Workbook()

# ── Styling constants ──────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBTLE_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBTLE_HEADER_FONT = Font(name="Calibri", bold=True, size=10)
CARD_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
CARD_TITLE_FONT = Font(name="Calibri", bold=False, size=10, color="666666")
CARD_VALUE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
INPUT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FORMULA_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
SECTION_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
LABEL_FONT = Font(name="Calibri", bold=True, size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

TAB_COLORS = {
    "Start Here": "2F5496",
    "Dashboard": "2E75B6",
    "Clients": "548235",
    "Invoice Register": "BF8F00",
    "Payment Tracker": "C00000",
    "Invoice Generator": "7030A0",
    "Settings": "404040",
}

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_formula_cell(cell):
    cell.fill = FORMULA_FILL
    cell.font = BODY_FONT
    cell.alignment = CENTER

def style_input_cell(cell):
    cell.fill = INPUT_FILL
    cell.font = BODY_FONT
    cell.alignment = LEFT
    cell.border = THIN_BORDER

def style_data_cell(cell, fmt=None):
    cell.font = BODY_FONT
    cell.alignment = LEFT
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt

def add_title(ws, row, col, title, merge_end_col=None):
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)

def add_section(ws, row, col, title, merge_end_col):
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
    for c in range(col, merge_end_col + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL
        ws.cell(row=row, column=c).font = SECTION_FONT

# ═════════════════════════════════════════════════════════════
# TAB 1: Start Here
# ═════════════════════════════════════════════════════════════
ws_start = wb.active
ws_start.title = "Start Here"
ws_start.sheet_properties.tabColor = TAB_COLORS["Start Here"]

ws_start.column_dimensions["A"].width = 80

add_title(ws_start, 1, 1, "Nivora Invoice Generator & Payment Tracker", 3)
ws_start.cell(row=2, column=1, value="Excel + Google Sheets compatible  |  Formula-driven  |  No macros").font = Font(name="Calibri", italic=True, size=10, color="666666")

row = 4
add_section(ws_start, row, 1, "Quick Start", 3)
row += 1
instructions = [
    "1. Go to the Settings tab to set your currency, default tax rate, and preferences.",
    "2. Add your clients on the Clients tab (or use the sample data as a guide).",
    "3. Create invoices using the Invoice Generator tab.",
    "4. Track payments on the Payment Tracker tab.",
    "5. View your dashboard for revenue, outstanding balances, and invoice status.",
]
for inst in instructions:
    ws_start.cell(row=row, column=1, value=inst).font = BODY_FONT
    ws_start.cell(row=row, column=1).alignment = LEFT
    row += 1

row += 1
add_section(ws_start, row, 1, "Tab Guide", 3)
row += 1
tab_guide = [
    ("Start Here", "Setup instructions and navigation guide"),
    ("Dashboard", "Revenue summary, invoice status breakdown, top clients"),
    ("Clients", "Manage client information and status"),
    ("Invoice Register", "All invoices with auto-status tracking"),
    ("Payment Tracker", "Log payments linked to invoices"),
    ("Invoice Generator", "Create printable invoices with auto-calculations"),
    ("Settings", "Configure currency, tax rate, dropdown values"),
]
ws_start.cell(row=row, column=1, value="Tab").font = LABEL_FONT
ws_start.cell(row=row, column=2, value="Purpose").font = LABEL_FONT
ws_start.cell(row=row, column=1).fill = SUBTLE_HEADER_FILL
ws_start.cell(row=row, column=2).fill = SUBTLE_HEADER_FILL
row += 1
for tab_name, purpose in tab_guide:
    ws_start.cell(row=row, column=1, value=tab_name).font = Font(name="Calibri", bold=True, size=10)
    ws_start.cell(row=row, column=2, value=purpose).font = BODY_FONT
    row += 1

row += 1
ws_start.cell(row=row, column=1, value="Important Notes").font = LABEL_FONT
row += 1
notes = [
    "Gray cells contain formulas — do not edit them.",
    "Green/white cells are for your data — enter values here.",
    "Sample data is included as an example. Replace it with your own data.",
    "Always keep a backup of your workbook.",
    "For Google Sheets: File > Save as Google Sheets. All formulas will convert automatically.",
    "Invoice Status is auto-calculated (Paid/Overdue/Partial/Sent). Set 'Draft' manually for work-in-progress.",
    "For support: workbysar1@gmail.com",
]
for note in notes:
    ws_start.cell(row=row, column=1, value=note).font = BODY_FONT
    ws_start.cell(row=row, column=1).alignment = LEFT
    row += 1

row += 1
ws_start.cell(row=row, column=1, value="© Nivora").font = Font(name="Calibri", italic=True, size=9, color="999999")

# ═════════════════════════════════════════════════════════════
# TAB 3: Clients (created before Dashboard for VLOOKUP refs)
# ═════════════════════════════════════════════════════════════
ws_clients = wb.create_sheet("Clients")
ws_clients.sheet_properties.tabColor = TAB_COLORS["Clients"]

client_headers = ["Client ID", "Client Name", "Company", "Email", "Phone", "Address", "Status", "Notes"]
col_widths_clients = [14, 22, 24, 30, 18, 30, 14, 30]
for i, w in enumerate(col_widths_clients, 1):
    ws_clients.column_dimensions[get_column_letter(i)].width = w

style_header_row(ws_clients, 1, len(client_headers))
for i, h in enumerate(client_headers, 1):
    ws_clients.cell(row=1, column=i, value=h)

clients_data = [
    ["CL-001", "Sarah Chen", "Chen Design Studio", "sarah@chendesign.com", "555-0101", "123 Design St, San Francisco, CA 94102", "Active", "Graphic design retainer"],
    ["CL-002", "Marcus Johnson", "Johnson Consulting", "marcus@johnsonconsulting.com", "555-0102", "456 Business Ave, Austin, TX 73301", "Active", "Monthly strategy consulting"],
    ["CL-003", "Elena Rodriguez", "Rodriguez Web Services", "elena@rodriguezweb.com", "555-0103", "789 Tech Blvd, Miami, FL 33101", "Active", "Web development projects"],
    ["CL-004", "David Kim", "Kim & Co. Marketing", "david@kimmarketing.com", "555-0104", "321 Market St, Chicago, IL 60601", "Active", "SEO and content marketing"],
    ["CL-005", "Aisha Patel", "Patel Digital Solutions", "aisha@pateldigital.com", "555-0105", "654 Innovation Dr, Seattle, WA 98101", "Active", "Digital transformation consulting"],
    ["CL-006", "James Wilson", "Wilson Creative Agency", "james@wilsoncreative.com", "555-0106", "987 Artisan Way, New York, NY 10001", "Active", "Brand design and strategy"],
    ["CL-007", "Priya Sharma", "Sharma Bookkeeping", "priya@sharmabooks.com", "555-0107", "147 Finance St, Denver, CO 80201", "Inactive", "Previous bookkeeping client"],
    ["CL-008", "Carlos Mendez", "Mendez Tech Services", "carlos@mendezteck.com", "555-0108", "258 Code Ln, Portland, OR 97201", "Active", "IT support retainer"],
    ["CL-009", "Olivia Thompson", "Thompson Freelance Writing", "olivia@thompsonwrite.com", "555-0109", "369 Wordsmith Rd, Boston, MA 02101", "Active", "Content writing projects"],
    ["CL-010", "Ben Okafor", "Okafor Design Lab", "ben@okaforlab.com", "555-0110", "741 Creative Ct, Atlanta, GA 30301", "Archived", "Completed project work"],
]

for r_idx, client in enumerate(clients_data, 2):
    for c_idx, val in enumerate(client, 1):
        cell = ws_clients.cell(row=r_idx, column=c_idx, value=val)
        cell.font = BODY_FONT
        cell.alignment = LEFT
        cell.border = THIN_BORDER
        if c_idx == 7:  # Status
            cell.alignment = CENTER

# Data validation for Client Status
dv_client_status = DataValidation(
    type="list",
    formula1='"Active,Inactive,Archived"',
    allow_blank=True,
)
dv_client_status.error = "Please select Active, Inactive, or Archived"
dv_client_status.errorTitle = "Invalid Status"
ws_clients.add_data_validation(dv_client_status)
dv_client_status.add(f"G2:G{len(clients_data)+1}")

ws_clients.freeze_panes = "A2"

# ═════════════════════════════════════════════════════════════
# TAB 4: Invoice Register (before Dashboard for formulas)
# ═════════════════════════════════════════════════════════════
ws_register = wb.create_sheet("Invoice Register")
ws_register.sheet_properties.tabColor = TAB_COLORS["Invoice Register"]

register_headers = ["Invoice ID", "Invoice Date", "Due Date", "Client ID", "Client Name", "Description", "Subtotal", "Tax Rate", "Tax Amount", "Total", "Amount Paid", "Outstanding", "Status"]
col_widths_reg = [14, 14, 14, 12, 22, 28, 14, 10, 14, 14, 14, 14, 14]
for i, w in enumerate(col_widths_reg, 1):
    ws_register.column_dimensions[get_column_letter(i)].width = w

style_header_row(ws_register, 1, len(register_headers))
for i, h in enumerate(register_headers, 1):
    ws_register.cell(row=1, column=i, value=h)

# Invoice data: (ID, Date, Due, ClientID, ClientName, Description, Subtotal, TaxRate, AmountPaid)
invoices_data = [
    ("INV-001", date(2025,1,5),  date(2025,2,5),  "CL-001", "Sarah Chen",       "Brand identity design",           1500, 0.10, 1650),
    ("INV-002", date(2025,1,8),  date(2025,2,8),  "CL-002", "Marcus Johnson",   "Q1 strategy consulting",          2500, 0.10, 2750),
    ("INV-003", date(2025,1,12), date(2025,2,12), "CL-003", "Elena Rodriguez",  "Website redesign",                 1800, 0.10, 1980),
    ("INV-004", date(2025,1,15), date(2025,2,15), "CL-004", "David Kim",        "SEO audit and recommendations",    3200, 0.10, 3520),
    ("INV-005", date(2025,1,20), date(2025,2,20), "CL-005", "Aisha Patel",      "Digital strategy workshop",        2100, 0.10, 2310),
    ("INV-006", date(2025,1,25), date(2025,2,25), "CL-006", "James Wilson",     "Brand campaign Q1",                4000, 0.10, 2200),
    ("INV-007", date(2025,2,2),  date(2025,3,2),  "CL-001", "Sarah Chen",       "Website UI design",                1600, 0.10, 1760),
    ("INV-008", date(2025,2,5),  date(2025,3,5),  "CL-008", "Carlos Mendez",    "IT infrastructure audit",           950, 0.10, 1045),
    ("INV-009", date(2025,2,10), date(2025,3,10), "CL-009", "Olivia Thompson",  "Blog content package",              750, 0.10, 825),
    ("INV-010", date(2025,2,12), date(2025,3,12), "CL-002", "Marcus Johnson",   "Market research report",           2800, 0.10, 1400),
    ("INV-011", date(2025,2,18), date(2025,3,18), "CL-003", "Elena Rodriguez",  "API integration",                  1500, 0.10, 0),
    ("INV-012", date(2025,2,22), date(2025,3,22), "CL-006", "James Wilson",     "Q2 campaign planning",             3500, 0.10, 0),
    ("INV-013", date(2025,2,25), date(2025,3,25), "CL-004", "David Kim",        "Monthly SEO retainer",             1800, 0.10, 0),
    ("INV-014", date(2025,3,1),  date(2025,4,1),  "CL-005", "Aisha Patel",      "Data analytics setup",             2200, 0.10, 1210),
    ("INV-015", date(2025,3,5),  date(2027,4,5),  "CL-001", "Sarah Chen",       "Mobile app mockups",               1400, 0.10, 1540),
    ("INV-016", date(2025,3,8),  date(2027,4,8),  "CL-008", "Carlos Mendez",    "Cloud migration support",          1200, 0.10, 1320),
    ("INV-017", date(2025,3,12), date(2027,4,12), "CL-009", "Olivia Thompson",  "Email newsletter setup",            600, 0.10, 0),
    ("INV-018", date(2025,3,15), date(2027,4,15), "CL-007", "Priya Sharma",     "Annual bookkeeping review",        3000, 0.10, 0),
    ("INV-019", date(2025,3,18), date(2025,4,18), "CL-006", "James Wilson",     "Brand guidelines document",        2800, 0.10, 0),
    ("INV-020", date(2025,3,22), date(2027,4,22), "CL-002", "Marcus Johnson",   "Q2 consulting services",           2000, 0.10, 1100),
    ("INV-021", date(2025,3,25), date(2027,4,25), "CL-003", "Elena Rodriguez",  "Performance optimization",         1700, 0.10, 0),
    ("INV-022", date(2025,3,28), date(2027,4,28), "CL-004", "David Kim",        "Content marketing plan",           1500, 0.10, 0),
    ("INV-023", date(2025,4,1),  date(2027,5,1),  "CL-005", "Aisha Patel",      "CRM implementation",               2600, 0.10, 1430),
    ("INV-024", date(2025,4,5),  date(2027,5,5),  "CL-001", "Sarah Chen",       "Product packaging design",         1800, 0.10, 0),
    ("INV-025", date(2025,4,8),  date(2027,5,8),  "CL-008", "Carlos Mendez",    "Server maintenance contract",       800, 0.10, 0),
]

for r_idx, inv in enumerate(invoices_data, 2):
    inv_id, inv_date, due_date, client_id, client_name, desc, subtotal, tax_rate, amt_paid = inv
    ws_register.cell(row=r_idx, column=1, value=inv_id).font = BODY_FONT
    ws_register.cell(row=r_idx, column=1).border = THIN_BORDER
    ws_register.cell(row=r_idx, column=1).alignment = CENTER

    ws_register.cell(row=r_idx, column=2, value=inv_date).font = BODY_FONT
    ws_register.cell(row=r_idx, column=2).number_format = "DD-MMM-YYYY"
    ws_register.cell(row=r_idx, column=2).border = THIN_BORDER
    ws_register.cell(row=r_idx, column=2).alignment = CENTER

    ws_register.cell(row=r_idx, column=3, value=due_date).font = BODY_FONT
    ws_register.cell(row=r_idx, column=3).number_format = "DD-MMM-YYYY"
    ws_register.cell(row=r_idx, column=3).border = THIN_BORDER
    ws_register.cell(row=r_idx, column=3).alignment = CENTER

    ws_register.cell(row=r_idx, column=4, value=client_id).font = BODY_FONT
    ws_register.cell(row=r_idx, column=4).border = THIN_BORDER
    ws_register.cell(row=r_idx, column=4).alignment = CENTER

    ws_register.cell(row=r_idx, column=5, value=client_name).font = BODY_FONT
    ws_register.cell(row=r_idx, column=5).border = THIN_BORDER
    ws_register.cell(row=r_idx, column=5).alignment = LEFT

    ws_register.cell(row=r_idx, column=6, value=desc).font = BODY_FONT
    ws_register.cell(row=r_idx, column=6).border = THIN_BORDER
    ws_register.cell(row=r_idx, column=6).alignment = LEFT

    ws_register.cell(row=r_idx, column=7, value=subtotal).font = BODY_FONT
    ws_register.cell(row=r_idx, column=7).number_format = '#,##0.00'
    ws_register.cell(row=r_idx, column=7).border = THIN_BORDER
    ws_register.cell(row=r_idx, column=7).alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")

    # Tax Rate (input)
    cell_tax_rate = ws_register.cell(row=r_idx, column=8, value=tax_rate)
    cell_tax_rate.font = BODY_FONT
    cell_tax_rate.number_format = '0%'
    cell_tax_rate.border = THIN_BORDER
    cell_tax_rate.alignment = CENTER
    cell_tax_rate.fill = INPUT_FILL

    # Tax Amount = Subtotal * Tax Rate  (formula - gray)
    cell_tax_amt = ws_register.cell(row=r_idx, column=9)
    cell_tax_amt.value = f"=G{r_idx}*H{r_idx}"
    style_formula_cell(cell_tax_amt)
    cell_tax_amt.number_format = '#,##0.00'

    # Total = Subtotal + Tax Amount (formula - gray)
    cell_total = ws_register.cell(row=r_idx, column=10)
    cell_total.value = f"=G{r_idx}+I{r_idx}"
    style_formula_cell(cell_total)
    cell_total.number_format = '#,##0.00'

    # Amount Paid (input)
    cell_amt_paid = ws_register.cell(row=r_idx, column=11, value=amt_paid)
    cell_amt_paid.font = BODY_FONT
    cell_amt_paid.number_format = '#,##0.00'
    cell_amt_paid.border = THIN_BORDER
    cell_amt_paid.alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
    cell_amt_paid.fill = INPUT_FILL

    # Outstanding = Total - Amount Paid (formula - gray)
    cell_outstanding = ws_register.cell(row=r_idx, column=12)
    cell_outstanding.value = f"=J{r_idx}-K{r_idx}"
    style_formula_cell(cell_outstanding)
    cell_outstanding.number_format = '#,##0.00'

    # Status formula (gray) — auto-produces Paid / Overdue / Partial / Sent
    # "Draft" is NOT auto-produced by this formula. Users can manually type "Draft"
    # to mark work-in-progress invoices; it is available as a dropdown value in
    # Settings for manual override.
    cell_status = ws_register.cell(row=r_idx, column=13)
    cell_status.value = f'=IF(L{r_idx}=0,"Paid",IF(TODAY()>C{r_idx},"Overdue",IF(K{r_idx}>0,"Partial","Sent")))'
    style_formula_cell(cell_status)

last_reg_row = len(invoices_data) + 1
ws_register.freeze_panes = "A2"

# ═════════════════════════════════════════════════════════════
# TAB 5: Payment Tracker
# ═════════════════════════════════════════════════════════════
ws_payments = wb.create_sheet("Payment Tracker")
ws_payments.sheet_properties.tabColor = TAB_COLORS["Payment Tracker"]

pay_headers = ["Payment ID", "Invoice ID", "Client Name", "Payment Date", "Payment Method", "Amount", "Notes"]
col_widths_pay = [14, 14, 24, 16, 20, 14, 30]
for i, w in enumerate(col_widths_pay, 1):
    ws_payments.column_dimensions[get_column_letter(i)].width = w

style_header_row(ws_payments, 1, len(pay_headers))
for i, h in enumerate(pay_headers, 1):
    ws_payments.cell(row=1, column=i, value=h)

payments_data = [
    ("PAY-001", "INV-001", date(2025,1,20),  "Bank Transfer", 1650, "Full payment"),
    ("PAY-002", "INV-002", date(2025,1,25),  "PayPal",        2750, "Invoice paid via PayPal"),
    ("PAY-003", "INV-003", date(2025,2,1),   "Bank Transfer", 1980, "Wire transfer"),
    ("PAY-004", "INV-004", date(2025,2,5),   "Cheque",        3520, "Cheque received"),
    ("PAY-005", "INV-005", date(2025,2,10),  "Bank Transfer", 2310, "Direct deposit"),
    ("PAY-006", "INV-006", date(2025,2,15),  "PayPal",        2200, "Partial payment - Q1 campaign"),
    ("PAY-007", "INV-007", date(2025,2,20),  "Bank Transfer", 1760, "Full payment"),
    ("PAY-008", "INV-008", date(2025,3,1),   "Cash",          1045, "Paid in cash"),
    ("PAY-009", "INV-009", date(2025,3,5),   "Bank Transfer", 825,  "Full payment"),
    ("PAY-010", "INV-010", date(2025,3,1),   "PayPal",        1400, "Partial - dispute on remaining"),
    ("PAY-011", "INV-015", date(2026,3,15),  "Bank Transfer", 1540, "Full payment received"),
    ("PAY-012", "INV-016", date(2026,3,20),  "Bank Transfer", 1320, "Full payment"),
    ("PAY-013", "INV-020", date(2026,4,1),   "Credit Card",   1100, "Partial payment"),
    ("PAY-014", "INV-023", date(2026,4,10),  "PayPal",        1430, "Partial payment"),
    ("PAY-015", "INV-014", date(2025,3,15),  "Bank Transfer", 1210, "Partial payment"),
]

for r_idx, pay in enumerate(payments_data, 2):
    pay_id, inv_id, pay_date, method, amount, notes = pay
    ws_payments.cell(row=r_idx, column=1, value=pay_id).font = BODY_FONT
    ws_payments.cell(row=r_idx, column=1).border = THIN_BORDER
    ws_payments.cell(row=r_idx, column=1).alignment = CENTER

    ws_payments.cell(row=r_idx, column=2, value=inv_id).font = BODY_FONT
    ws_payments.cell(row=r_idx, column=2).border = THIN_BORDER
    ws_payments.cell(row=r_idx, column=2).alignment = CENTER

    # Client Name (VLOOKUP formula from Invoice Register - gray formula cell)
    cell_client = ws_payments.cell(row=r_idx, column=3)
    cell_client.value = f"=IFERROR(VLOOKUP(B{r_idx},'Invoice Register'!A:E,5,FALSE),\"\")"
    style_formula_cell(cell_client)
    cell_client.border = THIN_BORDER

    ws_payments.cell(row=r_idx, column=4, value=pay_date).font = BODY_FONT
    ws_payments.cell(row=r_idx, column=4).number_format = "DD-MMM-YYYY"
    ws_payments.cell(row=r_idx, column=4).border = THIN_BORDER
    ws_payments.cell(row=r_idx, column=4).alignment = CENTER

    # Payment Method (input)
    cell_method = ws_payments.cell(row=r_idx, column=5, value=method)
    cell_method.font = BODY_FONT
    cell_method.border = THIN_BORDER
    cell_method.alignment = CENTER
    cell_method.fill = INPUT_FILL

    ws_payments.cell(row=r_idx, column=6, value=amount).font = BODY_FONT
    ws_payments.cell(row=r_idx, column=6).number_format = '#,##0.00'
    ws_payments.cell(row=r_idx, column=6).border = THIN_BORDER
    ws_payments.cell(row=r_idx, column=6).alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")

    ws_payments.cell(row=r_idx, column=7, value=notes).font = BODY_FONT
    ws_payments.cell(row=r_idx, column=7).border = THIN_BORDER
    ws_payments.cell(row=r_idx, column=7).alignment = LEFT

last_pay_row = len(payments_data) + 1
ws_payments.freeze_panes = "A2"

# Data validation for Payment Method
dv_payment_method = DataValidation(
    type="list",
    formula1='"Bank Transfer,PayPal,Cash,Cheque,Credit Card,Other"',
    allow_blank=True,
)
dv_payment_method.error = "Please select a valid payment method"
dv_payment_method.errorTitle = "Invalid Payment Method"
ws_payments.add_data_validation(dv_payment_method)
if last_pay_row > 1:
    dv_payment_method.add(f"E2:E{last_pay_row}")

# ═════════════════════════════════════════════════════════════
# TAB 2: Dashboard
# ═════════════════════════════════════════════════════════════
ws_dash = wb.create_sheet("Dashboard")
ws_dash.sheet_properties.tabColor = TAB_COLORS["Dashboard"]
ws_dash.sheet_view.showGridLines = False

# Column widths
for c in range(1, 9):
    ws_dash.column_dimensions[get_column_letter(c)].width = 18

add_title(ws_dash, 1, 1, "Nivora Invoice Generator & Payment Tracker", 8)

# ── KPI Cards Row 1 ──
kpi_row1 = 3
kpis1 = [
    ("Total Revenue", f"=SUMIFS('Invoice Register'!J:J,'Invoice Register'!M:M,\"<>Draft\")"),
    ("Paid Revenue", f"=SUMIFS('Invoice Register'!J:J,'Invoice Register'!M:M,\"Paid\")"),
    ("Outstanding Revenue", f"=SUMIFS('Invoice Register'!L:L,'Invoice Register'!M:M,\"Sent\",'Invoice Register'!L:L,\">0\")+SUMIFS('Invoice Register'!L:L,'Invoice Register'!M:M,\"Overdue\",'Invoice Register'!L:L,\">0\")+SUMIFS('Invoice Register'!L:L,'Invoice Register'!M:M,\"Partial\",'Invoice Register'!L:L,\">0\")"),
    ("Total Invoices", f"=COUNTIF('Invoice Register'!M:M,\"<>\")-1"),
]

for i, (label, formula) in enumerate(kpis1):
    col = 1 + i * 2
    c_label = ws_dash.cell(row=kpi_row1, column=col, value=label)
    c_label.font = CARD_TITLE_FONT
    c_label.fill = CARD_FILL
    c_label.alignment = CENTER
    c_label.border = THIN_BORDER
    if i < 3:
        ws_dash.merge_cells(start_row=kpi_row1, start_column=col, end_row=kpi_row1, end_column=col)
    else:
        ws_dash.merge_cells(start_row=kpi_row1, start_column=col, end_row=kpi_row1, end_column=col + 1)

    c_val = ws_dash.cell(row=kpi_row1 + 1, column=col)
    c_val.value = formula
    c_val.font = CARD_VALUE_FONT
    c_val.fill = CARD_FILL
    c_val.alignment = CENTER
    c_val.border = THIN_BORDER
    c_val.number_format = '#,##0.00' if label != "Total Invoices" else '#,##0'
    if i < 3:
        ws_dash.merge_cells(start_row=kpi_row1 + 1, start_column=col, end_row=kpi_row1 + 1, end_column=col)
    else:
        ws_dash.merge_cells(start_row=kpi_row1 + 1, start_column=col, end_row=kpi_row1 + 1, end_column=col + 1)

# ── KPI Cards Row 2 ──
kpi_row2 = 6
kpis2 = [
    ("Paid Invoices", f"=COUNTIFS('Invoice Register'!M:M,\"Paid\")", '#,##0'),
    ("Unpaid Invoices", f"=COUNTIFS('Invoice Register'!M:M,\"Sent\")+COUNTIFS('Invoice Register'!M:M,\"Overdue\")+COUNTIFS('Invoice Register'!M:M,\"Partial\")+COUNTIFS('Invoice Register'!M:M,\"Draft\")", '#,##0'),
    ("Overdue Invoices", f"=COUNTIFS('Invoice Register'!M:M,\"Overdue\")", '#,##0'),
    ("Active Clients", f"=COUNTIFS(Clients!G:G,\"Active\")", '#,##0'),
]

for i, (label, formula, fmt) in enumerate(kpis2):
    col = 1 + i * 2
    c_label = ws_dash.cell(row=kpi_row2, column=col, value=label)
    c_label.font = CARD_TITLE_FONT
    c_label.fill = CARD_FILL
    c_label.alignment = CENTER
    c_label.border = THIN_BORDER
    if i < 3:
        ws_dash.merge_cells(start_row=kpi_row2, start_column=col, end_row=kpi_row2, end_column=col)
    else:
        ws_dash.merge_cells(start_row=kpi_row2, start_column=col, end_row=kpi_row2, end_column=col + 1)

    c_val = ws_dash.cell(row=kpi_row2 + 1, column=col)
    c_val.value = formula
    c_val.font = CARD_VALUE_FONT
    c_val.fill = CARD_FILL
    c_val.alignment = CENTER
    c_val.border = THIN_BORDER
    c_val.number_format = fmt
    if i < 3:
        ws_dash.merge_cells(start_row=kpi_row2 + 1, start_column=col, end_row=kpi_row2 + 1, end_column=col)
    else:
        ws_dash.merge_cells(start_row=kpi_row2 + 1, start_column=col, end_row=kpi_row2 + 1, end_column=col + 1)

# Average Invoice Value
kpi_row3 = 9
c_avg_label = ws_dash.cell(row=kpi_row3, column=1, value="Average Invoice Value")
c_avg_label.font = CARD_TITLE_FONT
c_avg_label.fill = CARD_FILL
c_avg_label.alignment = CENTER
c_avg_label.border = THIN_BORDER
ws_dash.merge_cells(start_row=kpi_row3, start_column=1, end_row=kpi_row3, end_column=3)

c_avg_val = ws_dash.cell(row=kpi_row3 + 1, column=1)
c_avg_val.value = f"=IFERROR(SUMIFS('Invoice Register'!J2:J999,'Invoice Register'!M2:M999,\"<>Draft\")/COUNTIFS('Invoice Register'!M2:M999,\"<>Draft\",'Invoice Register'!M2:M999,\"<>\"),0)"
c_avg_val.font = CARD_VALUE_FONT
c_avg_val.fill = CARD_FILL
c_avg_val.alignment = CENTER
c_avg_val.border = THIN_BORDER
c_avg_val.number_format = '#,##0.00'
ws_dash.merge_cells(start_row=kpi_row3 + 1, start_column=1, end_row=kpi_row3 + 1, end_column=3)

# ── Revenue by Month ──
row = 13
add_section(ws_dash, row, 1, "Revenue by Month", 8)
row += 1

month_headers_r = ["Month", "Revenue"]
ws_dash.cell(row=row, column=1, value=month_headers_r[0]).font = LABEL_FONT
ws_dash.cell(row=row, column=1).fill = SUBTLE_HEADER_FILL
ws_dash.cell(row=row, column=1).border = THIN_BORDER
ws_dash.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws_dash.cell(row=row, column=4, value=month_headers_r[1]).font = LABEL_FONT
ws_dash.cell(row=row, column=4).fill = SUBTLE_HEADER_FILL
ws_dash.cell(row=row, column=4).border = THIN_BORDER
ws_dash.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)

months = [
    ("Jan 2025", 2025, 1),
    ("Feb 2025", 2025, 2),
    ("Mar 2025", 2025, 3),
    ("Apr 2025", 2025, 4),
]
for i, (label, yr, mo) in enumerate(months):
    r = row + 1 + i
    ws_dash.cell(row=r, column=1, value=label).font = BODY_FONT
    ws_dash.cell(row=r, column=1).alignment = CENTER
    ws_dash.cell(row=r, column=1).border = THIN_BORDER
    ws_dash.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

    cell_rev = ws_dash.cell(row=r, column=4)
    cell_rev.value = f"=SUMIFS('Invoice Register'!J:J,'Invoice Register'!M:M,\"<>Draft\",'Invoice Register'!B:B,\">=\"&DATE({yr},{mo},1),'Invoice Register'!B:B,\"<\"&DATE({yr},{mo}+1,1))"
    cell_rev.font = Font(name="Calibri", bold=True, size=11)
    cell_rev.number_format = '#,##0.00'
    cell_rev.alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
    cell_rev.border = THIN_BORDER
    ws_dash.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)

# ── Invoice Status Breakdown ──
row = 20
add_section(ws_dash, row, 1, "Invoice Status Breakdown", 8)
row += 1

status_h = ["Status", "Count"]
ws_dash.cell(row=row, column=1, value=status_h[0]).font = LABEL_FONT
ws_dash.cell(row=row, column=1).fill = SUBTLE_HEADER_FILL
ws_dash.cell(row=row, column=1).border = THIN_BORDER
ws_dash.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws_dash.cell(row=row, column=4, value=status_h[1]).font = LABEL_FONT
ws_dash.cell(row=row, column=4).fill = SUBTLE_HEADER_FILL
ws_dash.cell(row=row, column=4).border = THIN_BORDER
ws_dash.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)

statuses = ["Paid", "Sent", "Partial", "Overdue", "Draft"]
for i, st in enumerate(statuses):
    r = row + 1 + i
    ws_dash.cell(row=r, column=1, value=st).font = BODY_FONT
    ws_dash.cell(row=r, column=1).alignment = CENTER
    ws_dash.cell(row=r, column=1).border = THIN_BORDER
    ws_dash.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

    cell_cnt = ws_dash.cell(row=r, column=4)
    cell_cnt.value = f"=COUNTIFS('Invoice Register'!M:M,\"{st}\")"
    cell_cnt.font = Font(name="Calibri", bold=True, size=11)
    cell_cnt.number_format = '#,##0'
    cell_cnt.alignment = CENTER
    cell_cnt.border = THIN_BORDER
    ws_dash.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)

# ── Top Clients by Revenue ──
row = 28
add_section(ws_dash, row, 1, "Top Clients by Revenue", 8)
row += 1

top_h = ["Client", "Revenue"]
ws_dash.cell(row=row, column=1, value=top_h[0]).font = LABEL_FONT
ws_dash.cell(row=row, column=1).fill = SUBTLE_HEADER_FILL
ws_dash.cell(row=row, column=1).border = THIN_BORDER
ws_dash.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws_dash.cell(row=row, column=5, value=top_h[1]).font = LABEL_FONT
ws_dash.cell(row=row, column=5).fill = SUBTLE_HEADER_FILL
ws_dash.cell(row=row, column=5).border = THIN_BORDER
ws_dash.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)

top_client_names = ["James Wilson", "Sarah Chen", "Marcus Johnson", "Aisha Patel", "David Kim", "Elena Rodriguez"]
for i, cname in enumerate(top_client_names):
    r = row + 1 + i
    ws_dash.cell(row=r, column=1, value=cname).font = BODY_FONT
    ws_dash.cell(row=r, column=1).alignment = LEFT
    ws_dash.cell(row=r, column=1).border = THIN_BORDER
    ws_dash.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

    cell_cr = ws_dash.cell(row=r, column=5)
    cell_cr.value = f"=SUMIFS('Invoice Register'!J:J,'Invoice Register'!E:E,\"{cname}\",'Invoice Register'!M:M,\"<>Draft\")"
    cell_cr.font = Font(name="Calibri", bold=True, size=11)
    cell_cr.number_format = '#,##0.00'
    cell_cr.alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
    cell_cr.border = THIN_BORDER
    ws_dash.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)

# ═════════════════════════════════════════════════════════════
# TAB 6: Invoice Generator
# ═════════════════════════════════════════════════════════════
ws_gen = wb.create_sheet("Invoice Generator")
ws_gen.sheet_properties.tabColor = TAB_COLORS["Invoice Generator"]
ws_gen.sheet_view.showGridLines = False

for c in range(1, 9):
    ws_gen.column_dimensions[get_column_letter(c)].width = 18
ws_gen.column_dimensions["A"].width = 4
ws_gen.column_dimensions["B"].width = 32
ws_gen.column_dimensions["C"].width = 14
ws_gen.column_dimensions["D"].width = 14

# Title
ws_gen.cell(row=1, column=1, value="INVOICE").font = Font(name="Calibri", bold=True, size=24, color="2F5496")
ws_gen.merge_cells("A1:D1")

# Header info
row = 3
ws_gen.cell(row=row, column=1, value="Invoice Number:").font = LABEL_FONT
cell_inv_num = ws_gen.cell(row=row, column=2, value="INV-001")
cell_inv_num.fill = INPUT_FILL
cell_inv_num.font = BODY_FONT
cell_inv_num.border = THIN_BORDER

ws_gen.cell(row=row, column=3, value="Invoice Date:").font = LABEL_FONT
cell_inv_date = ws_gen.cell(row=row, column=4, value=date(2025, 6, 1))
cell_inv_date.fill = INPUT_FILL
cell_inv_date.font = BODY_FONT
cell_inv_date.number_format = "DD-MMM-YYYY"
cell_inv_date.border = THIN_BORDER

row = 4
ws_gen.cell(row=row, column=1, value="").font = LABEL_FONT
ws_gen.cell(row=row, column=3, value="Due Date:").font = LABEL_FONT
cell_due = ws_gen.cell(row=row, column=4, value=date(2025, 7, 1))
cell_due.fill = INPUT_FILL
cell_due.font = BODY_FONT
cell_due.number_format = "DD-MMM-YYYY"
cell_due.border = THIN_BORDER

# From section
row = 6
add_section(ws_gen, row, 1, "From", 4)
row += 1
ws_gen.cell(row=row, column=1, value="Your Name / Company:").font = LABEL_FONT
cell_from_name = ws_gen.cell(row=row, column=2, value="Your Company Name")
cell_from_name.fill = INPUT_FILL
cell_from_name.font = BODY_FONT
cell_from_name.border = THIN_BORDER
ws_gen.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

row += 1
ws_gen.cell(row=row, column=1, value="Your Address:").font = LABEL_FONT
cell_from_addr = ws_gen.cell(row=row, column=2, value="123 Business St, City, State ZIP")
cell_from_addr.fill = INPUT_FILL
cell_from_addr.font = BODY_FONT
cell_from_addr.border = THIN_BORDER
ws_gen.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

# Bill To section
row = 9
add_section(ws_gen, row, 1, "Bill To", 4)
row += 1
ws_gen.cell(row=row, column=1, value="Client Name:").font = LABEL_FONT
cell_client_name = ws_gen.cell(row=row, column=2, value="Sarah Chen")
cell_client_name.fill = INPUT_FILL
cell_client_name.font = BODY_FONT
cell_client_name.border = THIN_BORDER
ws_gen.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

row += 1
ws_gen.cell(row=row, column=1, value="Company:").font = LABEL_FONT
cell_company = ws_gen.cell(row=row, column=2, value="Chen Design Studio")
cell_company.fill = INPUT_FILL
cell_company.font = BODY_FONT
cell_company.border = THIN_BORDER
ws_gen.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

row += 1
ws_gen.cell(row=row, column=1, value="Address:").font = LABEL_FONT
cell_addr = ws_gen.cell(row=row, column=2, value="123 Design St, San Francisco, CA 94102")
cell_addr.fill = INPUT_FILL
cell_addr.font = BODY_FONT
cell_addr.border = THIN_BORDER
ws_gen.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

# Line items table
row = 14
add_section(ws_gen, row, 1, "Line Items", 4)
row += 1

line_headers = ["#", "Description", "Quantity", "Rate", "Amount"]
line_cols = [1, 2, 3, 4, 5]
ws_gen.column_dimensions["E"].width = 18

for i, h in enumerate(line_headers):
    col = line_cols[i]
    cell = ws_gen.cell(row=row, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

line_items = [
    (1, "Web design services", 40, 75),
    (2, "UI/UX consultation", 10, 150),
    (3, "Project management", 5, 100),
]

for i, (num, desc, qty, rate) in enumerate(line_items):
    r = row + 1 + i
    ws_gen.cell(row=r, column=1, value=num).font = BODY_FONT
    ws_gen.cell(row=r, column=1).alignment = CENTER
    ws_gen.cell(row=r, column=1).border = THIN_BORDER

    ws_gen.cell(row=r, column=2, value=desc).font = BODY_FONT
    ws_gen.cell(row=r, column=2).alignment = LEFT
    ws_gen.cell(row=r, column=2).border = THIN_BORDER

    cell_q = ws_gen.cell(row=r, column=3, value=qty)
    cell_q.fill = INPUT_FILL
    cell_q.font = BODY_FONT
    cell_q.alignment = CENTER
    cell_q.border = THIN_BORDER

    cell_r = ws_gen.cell(row=r, column=4, value=rate)
    cell_r.fill = INPUT_FILL
    cell_r.font = BODY_FONT
    cell_r.number_format = '#,##0.00'
    cell_r.alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
    cell_r.border = THIN_BORDER

    cell_a = ws_gen.cell(row=r, column=5)
    cell_a.value = f"=C{r}*D{r}"
    style_formula_cell(cell_a)
    cell_a.number_format = '#,##0.00'

# Empty rows for additional items
for i in range(3, 6):
    r = row + 1 + i
    for col in range(1, 6):
        cell = ws_gen.cell(row=r, column=col)
        cell.border = THIN_BORDER
        if col == 5:
            cell.value = f"=C{r}*D{r}"
            style_formula_cell(cell)
            cell.number_format = '#,##0.00'

# Summary section
summary_row = row + 1 + 6
add_section(ws_gen, summary_row, 1, "Summary", 5)
summary_row += 1

ws_gen.cell(row=summary_row, column=1, value="Subtotal:").font = LABEL_FONT
ws_gen.cell(row=summary_row, column=1).alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
ws_gen.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)
cell_sub = ws_gen.cell(row=summary_row, column=4)
cell_sub.value = f"=SUM(E{row+1}:E{row+5})"
cell_sub.font = Font(name="Calibri", bold=True, size=12)
cell_sub.number_format = '#,##0.00'
cell_sub.alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
ws_gen.merge_cells(start_row=summary_row, start_column=4, end_row=summary_row, end_column=5)

summary_row += 1
ws_gen.cell(row=summary_row, column=1, value="Tax Rate:").font = LABEL_FONT
ws_gen.cell(row=summary_row, column=1).alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
ws_gen.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)
cell_tr = ws_gen.cell(row=summary_row, column=4, value=0.10)
cell_tr.fill = INPUT_FILL
cell_tr.font = BODY_FONT
cell_tr.number_format = '0%'
cell_tr.alignment = CENTER
cell_tr.border = THIN_BORDER

summary_row += 1
ws_gen.cell(row=summary_row, column=1, value="Tax Amount:").font = LABEL_FONT
ws_gen.cell(row=summary_row, column=1).alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
ws_gen.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)
cell_ta = ws_gen.cell(row=summary_row, column=4)
cell_ta.value = f"=D{summary_row-1}*D{summary_row-2}"
cell_ta.font = Font(name="Calibri", bold=True, size=12)
cell_ta.number_format = '#,##0.00'
cell_ta.alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
ws_gen.merge_cells(start_row=summary_row, start_column=4, end_row=summary_row, end_column=5)

summary_row += 1
ws_gen.cell(row=summary_row, column=1, value="Total:").font = Font(name="Calibri", bold=True, size=13, color="2F5496")
ws_gen.cell(row=summary_row, column=1).alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
ws_gen.cell(row=summary_row, column=1).border = Border(top=Side(style="medium"), bottom=Side(style="medium"))
ws_gen.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)
cell_total = ws_gen.cell(row=summary_row, column=4)
cell_total.value = f"=D{summary_row-3}+D{summary_row-1}"
cell_total.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
cell_total.number_format = '#,##0.00'
cell_total.alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
cell_total.border = Border(top=Side(style="medium"), bottom=Side(style="medium"))
ws_gen.merge_cells(start_row=summary_row, start_column=4, end_row=summary_row, end_column=5)

# Payment Instructions
row = summary_row + 2
add_section(ws_gen, row, 1, "Payment Instructions", 5)
row += 1
ws_gen.cell(row=row, column=1, value="Payment Methods Accepted: Bank Transfer, PayPal, Credit Card").font = BODY_FONT
ws_gen.cell(row=row, column=1).alignment = LEFT_TOP
ws_gen.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

row += 1
ws_gen.cell(row=row, column=1, value="Payment Terms: Payment due within 30 days of invoice date.").font = BODY_FONT
ws_gen.cell(row=row, column=1).alignment = LEFT_TOP
ws_gen.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

# ═════════════════════════════════════════════════════════════
# TAB 7: Settings
# ═════════════════════════════════════════════════════════════
ws_settings = wb.create_sheet("Settings")
ws_settings.sheet_properties.tabColor = TAB_COLORS["Settings"]

col_widths_set = [30, 40]
for i, w in enumerate(col_widths_set, 1):
    ws_settings.column_dimensions[get_column_letter(i)].width = w

add_title(ws_settings, 1, 1, "Settings", 2)
ws_settings.cell(row=2, column=1, value="Edit values below to customize your workbook.").font = Font(name="Calibri", italic=True, size=10, color="666666")
ws_settings.merge_cells("A2:B2")

row = 4
add_section(ws_settings, row, 1, "General Settings", 2)
row += 1

ws_settings.cell(row=row, column=1, value="Currency Symbol").font = Font(name="Calibri", bold=True, size=10)
ws_settings.cell(row=row, column=1).fill = SUBTLE_HEADER_FILL
ws_settings.cell(row=row, column=1).border = THIN_BORDER
cell_currency = ws_settings.cell(row=row, column=2, value="$")
cell_currency.fill = INPUT_FILL
cell_currency.font = Font(name="Calibri", size=14, bold=True)
cell_currency.alignment = CENTER
cell_currency.border = THIN_BORDER

row += 1
ws_settings.cell(row=row, column=1, value="Default Tax Rate").font = Font(name="Calibri", bold=True, size=10)
ws_settings.cell(row=row, column=1).fill = SUBTLE_HEADER_FILL
ws_settings.cell(row=row, column=1).border = THIN_BORDER
cell_tax = ws_settings.cell(row=row, column=2, value=0.10)
cell_tax.fill = INPUT_FILL
cell_tax.font = BODY_FONT
cell_tax.number_format = '0%'
cell_tax.alignment = CENTER
cell_tax.border = THIN_BORDER

row += 2
add_section(ws_settings, row, 1, "Invoice Status Values", 2)
row += 1
ws_settings.cell(row=row, column=1, value="Status").font = LABEL_FONT
ws_settings.cell(row=row, column=1).fill = SUBTLE_HEADER_FILL
ws_settings.cell(row=row, column=1).border = THIN_BORDER
status_list = ["Sent", "Partial", "Paid", "Overdue", "Draft"]
for i, s in enumerate(status_list):
    r = row + 1 + i
    ws_settings.cell(row=r, column=1, value=s).font = BODY_FONT
    ws_settings.cell(row=r, column=1).alignment = CENTER
    ws_settings.cell(row=r, column=1).border = THIN_BORDER

row += len(status_list) + 2
add_section(ws_settings, row, 1, "Payment Method Values", 2)
row += 1
ws_settings.cell(row=row, column=1, value="Payment Method").font = LABEL_FONT
ws_settings.cell(row=row, column=1).fill = SUBTLE_HEADER_FILL
ws_settings.cell(row=row, column=1).border = THIN_BORDER
payment_methods = ["Bank Transfer", "PayPal", "Cash", "Cheque", "Credit Card", "Other"]
for i, p in enumerate(payment_methods):
    r = row + 1 + i
    ws_settings.cell(row=r, column=1, value=p).font = BODY_FONT
    ws_settings.cell(row=r, column=1).alignment = CENTER
    ws_settings.cell(row=r, column=1).border = THIN_BORDER

row += len(payment_methods) + 2
add_section(ws_settings, row, 1, "Client Status Values", 2)
row += 1
ws_settings.cell(row=row, column=1, value="Status").font = LABEL_FONT
ws_settings.cell(row=row, column=1).fill = SUBTLE_HEADER_FILL
ws_settings.cell(row=row, column=1).border = THIN_BORDER
client_statuses = ["Active", "Inactive", "Archived"]
for i, s in enumerate(client_statuses):
    r = row + 1 + i
    ws_settings.cell(row=r, column=1, value=s).font = BODY_FONT
    ws_settings.cell(row=r, column=1).alignment = CENTER
    ws_settings.cell(row=r, column=1).border = THIN_BORDER

# ═════════════════════════════════════════════════════════════
# Reorder sheets to match spec
# ═════════════════════════════════════════════════════════════
desired_order = ["Start Here", "Dashboard", "Clients", "Invoice Register", "Payment Tracker", "Invoice Generator", "Settings"]
for i, name in enumerate(desired_order):
    idx = wb.sheetnames.index(name)
    if idx != i:
        wb.move_sheet(name, offset=i - idx)

# ═════════════════════════════════════════════════════════════
# Save
# ═════════════════════════════════════════════════════════════
output_path = r"C:\Users\sar1s\Documents\MyWork\Git\digital-lab\01-product-factory\02-products\invoice-generator-payment-tracker\build\nivora-invoice-generator-payment-tracker.xlsx"
wb.save(output_path)
print(f"Workbook saved: {output_path}")
print(f"Tabs: {wb.sheetnames}")
