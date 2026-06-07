import openpyxl
import re
from collections import defaultdict

path = r"C:\Users\sar1s\Documents\MyWork\Git\digital-lab\01-product-factory\02-products\invoice-generator-payment-tracker\build\nivora-invoice-generator-payment-tracker.xlsx"
wb = openpyxl.load_workbook(path, data_only=False)

ALLOWED_FUNCTIONS = {
    "SUM", "SUMIF", "SUMIFS",
    "COUNTIF", "COUNTIFS",
    "IF", "IFERROR",
    "TODAY", "NOW",
    "VLOOKUP", "HLOOKUP",
    "TEXT",
    "AVERAGE", "AVERAGEIF", "AVERAGEIFS",
    "DATE", "YEAR", "MONTH", "DAY",
    "MIN", "MAX",
    "ROUND", "ROUNDUP", "ROUNDDOWN",
    "LEFT", "RIGHT", "MID", "LEN",
    "TRIM", "UPPER", "LOWER",
    "&",
}

EXCEL_ONLY_FUNCTIONS = {
    "XLOOKUP", "XMATCH",
    "FILTER", "SORT", "UNIQUE", "SEQUENCE", "SORTBY",
    "LET", "LAMBDA",
    "SWITCH",
    "TEXTJOIN",
    "CONCAT",
    "IFS",
    "MAXIFS", "MINIFS",
}

RISKY_GS_FUNCTIONS = {
    "INDIRECT",
    "OFFSET",
    "HYPERLINK",
    "GETPIVOTDATA",
}

all_pass = True
def check(ok, msg):
    global all_pass
    if ok:
        print(f"  PASS: {msg}")
    else:
        print(f"  FAIL: {msg}")
        all_pass = False

def warn(msg):
    print(f"  WARN: {msg}")

print("=" * 60)
print("  GOOGLE SHEETS COMPATIBILITY LINT")
print("=" * 60)

# ── Scan all formulas ──────────────────────────────────────────
all_formulas = []  # (sheet, cell, formula)
found_functions = set()
function_pattern = re.compile(r'([A-Z][A-Z0-9.]+)\(')

for sname in wb.sheetnames:
    ws = wb[sname]
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                all_formulas.append((sname, cell.coordinate, v))
                funcs = function_pattern.findall(v.upper())
                for f in funcs:
                    found_functions.add(f)

print(f"\n--- L-001: Total formulas scanned ---")
check(len(all_formulas) > 0, f"{len(all_formulas)} formulas found across {len(wb.sheetnames)} sheets")

print(f"\n--- L-002: Disallowed functions ---")
excel_only_used = found_functions & EXCEL_ONLY_FUNCTIONS
risky_used = found_functions & RISKY_GS_FUNCTIONS
check(len(excel_only_used) == 0, f"No Excel-only functions (found: {excel_only_used or 'none'})")
check(len(risky_used) == 0, f"No risky GS functions (found: {risky_used or 'none'})")

print(f"\n--- L-003: External workbook references ---")
ext_refs = []
for s, c, f in all_formulas:
    if "[" in f and "]" in f:
        ext_refs.append((s, c))
check(len(ext_refs) == 0, f"No external workbook references (found: {len(ext_refs)})")

print(f"\n--- L-004: Structured table references ---")
struct_refs = []
for s, c, f in all_formulas:
    if "@" in f or "[@" in f:
        struct_refs.append((s, c, f))
check(len(struct_refs) == 0, f"No structured table references (found: {len(struct_refs)})")

print(f"\n--- L-005: Dynamic array spill operators ---")
spill_refs = []
for s, c, f in all_formulas:
    if "#" in f and not any(ign in f for ign in ["<>", "#N/A"]):
        # Check if # is used as spill operator (not comparison)
        if re.search(r'[A-Z]#', f) or re.search(r'#REF', f):
            spill_refs.append((s, c, f))
check(len(spill_refs) == 0, f"No spill/reference errors (found: {len(spill_refs)})")

print(f"\n--- L-006: Local file path references ---")
path_refs = []
for s, c, f in all_formulas:
    if "\\\\" in f or "C:" in f or "D:" in f or "/Users/" in f:
        path_refs.append((s, c, f))
check(len(path_refs) == 0, f"No local path references (found: {len(path_refs)})")

print(f"\n--- L-007: VBA / macro references ---")
check(wb.vba_archive is None, "No VBA project present")
# Also check for any FORMULA in cells that reference macros
macro_refs = []
for s, c, f in all_formulas:
    if "EXEC" in f.upper() or "CALL" in f.upper() and "(" in f:
        macro_refs.append((s, c, f))
check(len(macro_refs) == 0, f"No macro references in formulas (found: {len(macro_refs)})")

print(f"\n--- L-008: Functions used (allowed) ---")
allowed_in_use = found_functions & ALLOWED_FUNCTIONS
if allowed_in_use:
    print(f"  Functions used: {sorted(allowed_in_use)}")
unknown = found_functions - ALLOWED_FUNCTIONS - EXCEL_ONLY_FUNCTIONS - RISKY_GS_FUNCTIONS
if unknown:
    for fn in sorted(unknown):
        warn(f"Unknown function: {fn}")

print(f"\n--- L-009: Sheet name quoting ---")
unquoted_refs = []
for s, c, f in all_formulas:
    refs = re.findall(r"(\w[\w ]*\w)!", f)
    for ref in refs:
        if " " in ref:
            unquoted_refs.append((s, c, ref))
check(len(unquoted_refs) == 0, f"All multi-word sheet names are single-quoted (found: {len(unquoted_refs)})")

# ── Per-sheet formula count ────────────────────────────────────
print(f"\n--- L-010: Formula distribution ---")
sheets_with_formulas = defaultdict(int)
for s, c, f in all_formulas:
    sheets_with_formulas[s] += 1
for sname in wb.sheetnames:
    cnt = sheets_with_formulas.get(sname, 0)
    print(f"  {sname}: {cnt} formulas")

# ── L-011: Data validation sources ─────────────────────────────
print(f"\n--- L-011: Data validation sources ---")
dv_ok = True
for ws in wb.worksheets:
    for dv in ws.data_validations.dataValidation:
        if dv.formula1 and not dv.formula1.startswith('"') and "!" in dv.formula1:
            dv_ok = False
            check(False, f"{ws.title}: DV references range {dv.formula1} — use inline list")
if dv_ok:
    check(True, "All data validations use inline lists (GS-compatible)")

# ── Summary ────────────────────────────────────────────────────
print(f"\n{'='*60}")
if all_pass:
    print("ALL GOOGLE SHEETS COMPATIBILITY CHECKS PASSED")
else:
    print("SOME GOOGLE SHEETS COMPATIBILITY CHECKS FAILED")
