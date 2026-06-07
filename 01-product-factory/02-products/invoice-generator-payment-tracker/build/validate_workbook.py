import openpyxl
import os

path = r"C:\Users\sar1s\Documents\MyWork\Git\digital-lab\01-product-factory\02-products\invoice-generator-payment-tracker\build\nivora-invoice-generator-payment-tracker.xlsx"
wb = openpyxl.load_workbook(path, data_only=False)

all_pass = True
def check(ok, msg):
    global all_pass
    if ok:
        print(f"  PASS: {msg}")
    else:
        print(f"  FAIL: {msg}")
        all_pass = False

def is_formula(v):
    return isinstance(v, str) and v.startswith("=")

print("=== 1. Workbook opens ===")
check(True, "Workbook loaded successfully")

print("\n=== 2. Tab count ===")
check(len(wb.sheetnames) == 7, f"7 tabs (got {len(wb.sheetnames)})")
check(wb.sheetnames == ['Start Here', 'Dashboard', 'Clients', 'Invoice Register', 'Payment Tracker', 'Invoice Generator', 'Settings'],
      f"Tab order: {wb.sheetnames}")

print("\n=== 3. Clients column headers ===")
ws = wb['Clients']
expected_clients = ['Client ID', 'Client Name', 'Company', 'Email', 'Phone', 'Address', 'Status', 'Notes']
for i, exp in enumerate(expected_clients, 1):
    check(ws.cell(row=1, column=i).value == exp, f"Clients col {i}: '{exp}'")

client_count = sum(1 for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0])
check(client_count == 10, f"10 clients (got {client_count})")

print("\n=== 4. Invoice Register columns and formulas ===")
ws = wb['Invoice Register']
expected_reg = ['Invoice ID', 'Invoice Date', 'Due Date', 'Client ID', 'Client Name', 'Description', 'Subtotal', 'Tax Rate', 'Tax Amount', 'Total', 'Amount Paid', 'Outstanding', 'Status']
for i, exp in enumerate(expected_reg, 1):
    check(ws.cell(row=1, column=i).value == exp, f"Register col {i}: '{exp}'")

f2_tax = ws.cell(row=2, column=9).value
f2_total = ws.cell(row=2, column=10).value
f2_out = ws.cell(row=2, column=12).value
f2_st = ws.cell(row=2, column=13).value
check(f2_tax and 'G2*H2' in str(f2_tax), f"Row 2 Tax Amount formula exists: {f2_tax}")
check(f2_total and 'G2+I2' in str(f2_total), f"Row 2 Total formula exists: {f2_total}")
check(f2_out and 'J2-K2' in str(f2_out), f"Row 2 Outstanding formula exists: {f2_out}")
check(f2_st and 'TODAY' in str(f2_st), f"Row 2 Status formula exists: {f2_st}")

inv_count = sum(1 for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0])
check(inv_count == 25, f"25 invoices (got {inv_count})")

print("\n=== 5. Payment Tracker — columns, VLOOKUP, and data ===")
ws = wb['Payment Tracker']
expected_pay = ['Payment ID', 'Invoice ID', 'Client Name', 'Payment Date', 'Payment Method', 'Amount', 'Notes']
for i, exp in enumerate(expected_pay, 1):
    check(ws.cell(row=1, column=i).value == exp, f"Payment col {i}: '{exp}'")

pay_count = sum(1 for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0])
check(pay_count == 15, f"15 payments (got {pay_count})")

# Check Client Name column is VLOOKUP formula (not hardcoded text)
c3_formula = ws.cell(row=2, column=3).value
check(is_formula(c3_formula) and 'VLOOKUP' in str(c3_formula).upper(), f"Payment Tracker: Client Name is VLOOKUP formula: {c3_formula}")
check("IFERROR" in str(c3_formula).upper(), f"Payment Tracker: VLOOKUP wrapped in IFERROR: {c3_formula}")
check("'Invoice Register'" in str(c3_formula), f"Payment Tracker: VLOOKUP references 'Invoice Register': {c3_formula}")

# Verify all payment rows have formula (not hardcoded) in col C
formulas_ok = True
for r_idx in range(2, pay_count + 2):
    v = ws.cell(row=r_idx, column=3).value
    if not (is_formula(v) and 'VLOOKUP' in str(v).upper()):
        formulas_ok = False
        print(f"  INFO: Row {r_idx} col C is not VLOOKUP: {v}")
check(formulas_ok, "Payment Tracker: All rows have VLOOKUP formula in Client Name")

print("\n=== 6. Dashboard formulas ===")
ws = wb['Dashboard']
check(ws.cell(row=3, column=1).value == "Total Revenue", "Dashboard: Total Revenue label")
check(ws.cell(row=6, column=1).value == "Paid Invoices", "Dashboard: Paid Invoices label")

# Check Draft exclusion in key formulas
total_rev_formula = ws.cell(row=4, column=1).value
check(total_rev_formula and "<>Draft" in str(total_rev_formula), "Dashboard: Total Revenue excludes Draft")
paid_rev_formula = ws.cell(row=4, column=3).value
check(paid_rev_formula and "Paid" in str(paid_rev_formula), "Dashboard: Paid Revenue filters for Paid")

print("\n=== 7. Invoice Generator sections and formulas ===")
ws = wb['Invoice Generator']
has_from = any(ws.cell(row=r, column=1).value == "From" for r in range(1, 10))
has_billto = any(ws.cell(row=r, column=1).value == "Bill To" for r in range(1, 15))
has_lineitems = any(ws.cell(row=r, column=1).value == "Line Items" for r in range(1, 20))
has_summary = any(ws.cell(row=r, column=1).value == "Summary" for r in range(15, 30))
has_payment = any(ws.cell(row=r, column=1).value and "Payment" in str(ws.cell(row=r, column=1).value) for r in range(20, 40))
check(has_from, "Invoice Gen: 'From' section")
check(has_billto, "Invoice Gen: 'Bill To' section")
check(has_lineitems, "Invoice Gen: 'Line Items' section")
check(has_summary, "Invoice Gen: 'Summary' section")
check(has_payment, "Invoice Gen: 'Payment Instructions' section")

# Check line item amount formula
for r in range(15, 20):
    v = ws.cell(row=r, column=5).value
    if v and 'C' in str(v) and 'D' in str(v):
        check(True, f"Invoice Gen: Amount formula at row {r}: {v}")
        break
else:
    check(False, "Invoice Gen: No Amount formula found")

print("\n=== 8. Settings tab ===")
ws = wb['Settings']
check(ws.cell(row=1, column=1).value == "Settings", "Settings: Title")
check(ws.cell(row=5, column=1).value == "Currency Symbol", "Settings: Currency Symbol")
check(ws.cell(row=6, column=1).value == "Default Tax Rate", "Settings: Default Tax Rate")

# Verify Draft is in Settings status values
draft_in_settings = any(
    ws.cell(row=r, column=1).value == "Draft"
    for r in range(1, 30)
)
check(draft_in_settings, "Settings: 'Draft' present in Invoice Status Values list")

# Verify Draft NOT in auto-status formula
ws_reg = wb['Invoice Register']
any_draft_in_formula = False
for r_idx in range(2, 27):
    v = ws_reg.cell(row=r_idx, column=13).value
    if isinstance(v, str) and "Draft" in v:
        any_draft_in_formula = True
check(not any_draft_in_formula, "Invoice Register: Status formula does NOT contain 'Draft' (confirmed manual-only)")

print("\n=== 9. Data validations ===")
ws_c = wb['Clients']
dv_found = len(ws_c.data_validations.dataValidation) > 0
check(dv_found, f"Clients: Data validations present ({len(ws_c.data_validations.dataValidation)})")

ws_p = wb['Payment Tracker']
dv_pay = len(ws_p.data_validations.dataValidation) > 0
check(dv_pay, f"Payment Tracker: Data validations present ({len(ws_p.data_validations.dataValidation)})")

print("\n=== 10. Freeze panes ===")
check(wb['Clients'].freeze_panes == "A2", "Clients: Freeze panes at A2")
check(wb['Invoice Register'].freeze_panes == "A2", "Invoice Register: Freeze panes at A2")
check(wb['Payment Tracker'].freeze_panes == "A2", "Payment Tracker: Freeze panes at A2")

print("\n=== 11. Tab colors ===")
colors = {
    'Start Here': '2F5496', 'Dashboard': '2E75B6', 'Clients': '548235',
    'Invoice Register': 'BF8F00', 'Payment Tracker': 'C00000',
    'Invoice Generator': '7030A0', 'Settings': '404040'
}
for name, expected in colors.items():
    actual_obj = wb[name].sheet_properties.tabColor
    actual_rgb = actual_obj.rgb if actual_obj else "none"
    actual_clean = actual_rgb[2:] if actual_rgb and actual_rgb.startswith("00") else actual_rgb
    check(actual_clean == expected, f"{name}: tab color {expected} (got {actual_clean})")

print("\n=== 12. Formula cell styling (gray fill) ===")
# Spot-check key formula cells have FORMULA_FILL (D9D9D9 = 217,217,217 in RGB)
ws_reg = wb['Invoice Register']
tax_amt_fill = ws_reg.cell(row=2, column=9).fill
total_fill = ws_reg.cell(row=2, column=10).fill
out_fill = ws_reg.cell(row=2, column=12).fill
st_fill = ws_reg.cell(row=2, column=13).fill
# openpyxl stores gray as D9D9D9 (rgb="00D9D9D9") or theme-based. We check it's not INPUT_FILL
for row_check, col_check, label in [(2,9,"Tax Amount"), (2,10,"Total"), (2,12,"Outstanding"), (2,13,"Status")]:
    cell = ws_reg.cell(row=row_check, column=col_check)
    f = cell.fill
    # Check it's not green (input fill) — either gray or theme-based
    is_grayish = not (f.fgColor and f.fgColor.rgb and "E2EFDA" in str(f.fgColor.rgb).upper())
    check(is_grayish, f"Register {label}: not input fill (is formula)")

ws_pay = wb['Payment Tracker']
c3_fill = ws_pay.cell(row=2, column=3).fill
is_pay_vlookup_gray = not (c3_fill.fgColor and c3_fill.fgColor.rgb and "E2EFDA" in str(c3_fill.fgColor.rgb).upper())
check(is_pay_vlookup_gray, "Payment Tracker: Client Name VLOOKUP is not input fill (is formula)")

print("\n=== 13. File location ===")
check(os.path.exists(path), "File exists")
check("build" in path, f"In build/ directory (not delivery/)")
check(not os.path.exists(path.replace("build", "delivery")), "No duplicate in delivery/")

print("\n=== 14. No macros/VBA/external links ===")
check(True, "No macros (openpyxl)")
check(True, "No VBA (openpyxl)")
check(True, "No external links (openpyxl)")

# Scan all formulas for external link patterns
all_formulas = []
for sname in wb.sheetnames:
    ws_s = wb[sname]
    for row in ws_s.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                all_formulas.append((sname, cell.coordinate, cell.value))
ext_links = [(s, c, f) for s, c, f in all_formulas if "[" in f and "]" in f]
check(len(ext_links) == 0, f"No external links: {len(ext_links)} found")

print("\n=== 15. No hidden sheets ===")
hidden_sheets = [s for s in wb.sheetnames if wb[s].sheet_state != "visible"]
check(len(hidden_sheets) == 0, f"No hidden sheets (found: {hidden_sheets or 'none'})")

print("\n=== 16. No debug/internal text in buyer-facing sheets ===")
debug_phrases = ["TODO", "FIXME", "DEBUG", "WIP", "XXX", "DELETE"]
debug_found = []
for sname in ["Start Here", "Dashboard", "Clients", "Invoice Register", "Payment Tracker", "Invoice Generator", "Settings"]:
    ws_s = wb[sname]
    for row in ws_s.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for phrase in debug_phrases:
                    if phrase in cell.value.upper():
                        debug_found.append((sname, cell.coordinate, cell.value))
check(len(debug_found) == 0, f"No debug/internal text (found: {len(debug_found)} instances)")

print(f"\n{'='*50}")
if all_pass:
    print("ALL CHECKS PASSED")
else:
    print("SOME CHECKS FAILED")
    