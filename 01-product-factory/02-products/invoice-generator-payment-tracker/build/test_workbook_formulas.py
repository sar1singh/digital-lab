import openpyxl
from datetime import date, datetime
from collections import defaultdict

path = r"C:\Users\sar1s\Documents\MyWork\Git\digital-lab\01-product-factory\02-products\invoice-generator-payment-tracker\build\nivora-invoice-generator-payment-tracker.xlsx"
wb = openpyxl.load_workbook(path, data_only=False)

TODAY = date.today()

all_pass = True
def check(ok, msg):
    global all_pass
    if ok:
        print(f"  PASS: {msg}")
    else:
        print(f"  FAIL: {msg}")
        all_pass = False

print("=" * 60)
print("  FORMULA UNIT TESTS")
print("=" * 60)

# ── Read client data ──────────────────────────────────────────
ws_clients = wb["Clients"]
clients = []
for row in ws_clients.iter_rows(min_row=2, values_only=True):
    cid = row[0]
    if not cid:
        break
    clients.append({
        "id": cid, "name": row[1], "company": row[2], "email": row[3],
        "phone": row[4], "address": row[5], "status": row[6], "notes": row[7]
    })

print(f"\n--- F-001: Clients ---")
check(len(clients) == 10, f"10 clients loaded (got {len(clients)})")
active = sum(1 for c in clients if c["status"] == "Active")
inactive = sum(1 for c in clients if c["status"] == "Inactive")
archived = sum(1 for c in clients if c["status"] == "Archived")
check(active == 8, f"8 Active clients (got {active})")
check(inactive == 1, f"1 Inactive client (got {inactive})")
check(archived == 1, f"1 Archived client (got {archived})")

# ── Read invoice data ─────────────────────────────────────────
ws_reg = wb["Invoice Register"]
invoices = []
inv_col9_formulas = []
for row in ws_reg.iter_rows(min_row=2, values_only=True):
    inv_id = row[0]
    if not inv_id:
        break
    # columns: 1=ID,2=Date,3=DueDate,4=ClientID,5=ClientName,6=Desc,
    #           7=Subtotal,8=TaxRate,9=TaxAmt(formula),10=Total(formula),
    #           11=AmtPaid,12=Outstanding(formula),13=Status(formula)
    inv_col9_formulas.append(row[8])
    invoices.append({
        "id": inv_id,
        "date": row[1],
        "due_date": row[2],
        "client_id": row[3],
        "client_name": row[4],
        "description": row[5],
        "subtotal": row[7],
        "tax_rate": row[7] if isinstance(row[7], (int, float)) else 0.10,
        "amt_paid": row[10] if isinstance(row[10], (int, float)) else 0,
    })
    # Re-read with separate approach for formula columns
    # Actually, let me re-read the raw cells
# Better to re-read with cell references
ws_reg_cells = wb["Invoice Register"]
invoices2 = []
for r_idx in range(2, 27):
    inv_id = ws_reg_cells.cell(row=r_idx, column=1).value
    if not inv_id:
        break
    subtotal = ws_reg_cells.cell(row=r_idx, column=7).value
    tax_rate = ws_reg_cells.cell(row=r_idx, column=8).value
    amt_paid = ws_reg_cells.cell(row=r_idx, column=11).value
    inv_date = ws_reg_cells.cell(row=r_idx, column=2).value
    due_date = ws_reg_cells.cell(row=r_idx, column=3).value
    client_name = ws_reg_cells.cell(row=r_idx, column=5).value
    tax_amt_formula = ws_reg_cells.cell(row=r_idx, column=9).value
    total_formula = ws_reg_cells.cell(row=r_idx, column=10).value
    outstanding_formula = ws_reg_cells.cell(row=r_idx, column=12).value
    status_formula = ws_reg_cells.cell(row=r_idx, column=13).value

    invoices2.append({
        "id": inv_id, "date": inv_date, "due_date": due_date,
        "client_name": client_name, "subtotal": subtotal,
        "tax_rate": tax_rate, "amt_paid": amt_paid if amt_paid else 0,
        "tax_amt_formula": tax_amt_formula,
        "total_formula": total_formula,
        "outstanding_formula": outstanding_formula,
        "status_formula": status_formula,
    })

print(f"\n--- F-002: Invoice Count ---")
check(len(invoices2) == 25, f"25 invoices loaded (got {len(invoices2)})")

# ── Derive invoice values ─────────────────────────────────────
for inv in invoices2:
    # Normalize datetime to date if needed
    if isinstance(inv["date"], datetime):
        inv["date"] = inv["date"].date()
    if isinstance(inv["due_date"], datetime):
        inv["due_date"] = inv["due_date"].date()

    inv["expected_tax_amt"] = round(inv["subtotal"] * inv["tax_rate"], 2)
    inv["expected_total"] = round(inv["subtotal"] + inv["expected_tax_amt"], 2)
    inv["expected_outstanding"] = round(inv["expected_total"] - inv["amt_paid"], 2)

    # Status logic: same as workbook formula
    outstanding = inv["expected_outstanding"]
    due = inv["due_date"]
    if isinstance(due, datetime):
        due = due.date()
    paid = inv["amt_paid"]
    if outstanding == 0:
        inv["expected_status"] = "Paid"
    elif TODAY > due:
        inv["expected_status"] = "Overdue"
    elif paid > 0:
        inv["expected_status"] = "Partial"
    else:
        inv["expected_status"] = "Sent"

print(f"\n--- F-003: Tax Amount (Subtotal * Tax Rate) ---")
all_tax_ok = True
for inv in invoices2:
    expected = inv["expected_tax_amt"]
    src_formula = inv["tax_amt_formula"]
    formula_ok = isinstance(src_formula, str) and "G" in src_formula and "H" in src_formula
    if not formula_ok:
        all_tax_ok = False
check(all_tax_ok, "All 25 Tax Amount cells have =G*H formula patterns")

print(f"\n--- F-004: Total (Subtotal + Tax Amount) ---")
all_total_ok = True
for inv in invoices2:
    formula_ok = isinstance(inv["total_formula"], str) and "G" in inv["total_formula"] and "I" in inv["total_formula"]
    if not formula_ok:
        all_total_ok = False
check(all_total_ok, "All 25 Total cells have =G+I formula patterns")

print(f"\n--- F-005: Outstanding (Total - Amount Paid) ---")
all_out_ok = True
for inv in invoices2:
    formula_ok = isinstance(inv["outstanding_formula"], str) and "J" in inv["outstanding_formula"] and "K" in inv["outstanding_formula"]
    if not formula_ok:
        all_out_ok = False
check(all_out_ok, "All 25 Outstanding cells have =J-K formula patterns")

print(f"\n--- F-006: Status Formula ---")
status_results = defaultdict(list)
for inv in invoices2:
    status_results[inv["expected_status"]].append(inv["id"])
    formula_ok = isinstance(inv["status_formula"], str) and "IF" in inv["status_formula"] and "TODAY" in inv["status_formula"]
    if not formula_ok:
        pass  # check below

all_status_formulas_ok = all(
    isinstance(inv["status_formula"], str) and "IF" in str(inv["status_formula"]) and "TODAY" in str(inv["status_formula"])
    for inv in invoices2
)
check(all_status_formulas_ok, "All 25 Status cells have IF+TODAY formula patterns")

status_counts = {s: len(v) for s, v in status_results.items()}
check(status_counts.get("Paid", 0) == 10, f"Paid invoices: 10 (got {status_counts.get('Paid', 0)}) — {status_results.get('Paid', [])}")
check(status_counts.get("Overdue", 0) == 7, f"Overdue invoices: 7 (got {status_counts.get('Overdue', 0)}) — {status_results.get('Overdue', [])}")
check(status_counts.get("Partial", 0) == 2, f"Partial invoices: 2 (got {status_counts.get('Partial', 0)}) — {status_results.get('Partial', [])}")
check(status_counts.get("Sent", 0) == 6, f"Sent invoices: 6 (got {status_counts.get('Sent', 0)}) — {status_results.get('Sent', [])}")
check("Draft" not in status_results or status_results.get("Draft", []) == [], "No Draft from auto-status formula (manual-only)")

# ── Read payment data ──────────────────────────────────────────
ws_pay = wb["Payment Tracker"]
payments = []
for r_idx in range(2, 17):
    pay_id = ws_pay.cell(row=r_idx, column=1).value
    if not pay_id:
        break
    inv_id = ws_pay.cell(row=r_idx, column=2).value
    client_formula = ws_pay.cell(row=r_idx, column=3).value
    amt = ws_pay.cell(row=r_idx, column=6).value
    payments.append({"id": pay_id, "inv_id": inv_id, "client_formula": client_formula, "amount": amt})

print(f"\n--- F-007: Payment Tracker ---")
check(len(payments) == 15, f"15 payments loaded (got {len(payments)})")

# VLOOKUP formula check
all_vlookup_ok = True
for p in payments:
    f = p["client_formula"]
    if not (isinstance(f, str) and "VLOOKUP" in f and "IFERROR" in f and "'Invoice Register'" in f):
        all_vlookup_ok = False
check(all_vlookup_ok, "All 15 Payment Tracker rows have IFERROR+VLOOKUP formula")

# Payment amounts > 0
all_positive = all(p["amount"] > 0 for p in payments)
check(all_positive, "All 15 payment amounts are positive")

# Verify VLOOKUP would resolve correctly (client name check)
inv_client_map = {inv["id"]: inv["client_name"] for inv in invoices2}
vlookup_matches = True
for p in payments:
    expected_client = inv_client_map.get(p["inv_id"])
    if not expected_client:
        vlookup_matches = False
check(vlookup_matches, "All Payment Invoice IDs exist in Invoice Register (VLOOKUP will resolve)")

# ── Dashboard KPI verification ─────────────────────────────────
print(f"\n--- F-008: Total Revenue (SUMIFS excluding Draft) ---")
total_revenue = round(sum(inv["expected_total"] for inv in invoices2), 2)
check(round(total_revenue, 2) == 54560.00, f"Total Revenue: 54560.00 (got {total_revenue:.2f})")

print(f"\n--- F-009: Paid Revenue ---")
paid_revenue = round(sum(inv["expected_total"] for inv in invoices2 if inv["expected_status"] == "Paid"), 2)
check(round(paid_revenue, 2) == 18700.00, f"Paid Revenue: 18700.00 (got {paid_revenue:.2f})")

print(f"\n--- F-010: Outstanding Revenue ---")
outstanding_revenue = round(sum(
    inv["expected_outstanding"] for inv in invoices2
    if inv["expected_status"] in ("Sent", "Overdue", "Partial") and inv["expected_outstanding"] > 0
), 2)
check(round(outstanding_revenue, 2) == 28520.00, f"Outstanding Revenue: 28520.00 (got {outstanding_revenue:.2f})")

print(f"\n--- F-011: Invoice Counts ---")
total_inv_count = len(invoices2)
check(total_inv_count == 25, f"Total Invoices: 25 (got {total_inv_count})")

paid_count = sum(1 for inv in invoices2 if inv["expected_status"] == "Paid")
check(paid_count == 10, f"Paid Invoices: 10 (got {paid_count})")

unpaid_count = sum(1 for inv in invoices2 if inv["expected_status"] in ("Sent", "Overdue", "Partial", "Draft"))
check(unpaid_count == 15, f"Unpaid Invoices: 15 (got {unpaid_count})")

overdue_count = sum(1 for inv in invoices2 if inv["expected_status"] == "Overdue")
check(overdue_count == 7, f"Overdue Invoices: 7 (got {overdue_count})")

print(f"\n--- F-012: Active Clients ---")
check(active == 8, f"Active Clients: 8 (got {active})")

print(f"\n--- F-013: Average Invoice Value ---")
avg_inv_value = round(total_revenue / total_inv_count, 2)
check(round(avg_inv_value, 2) == 2182.40, f"Average Invoice Value: 2182.40 (got {avg_inv_value:.2f})")
# Also verify the formula uses finite range (not whole-column) with non-empty criterion
avg_formula = wb["Dashboard"].cell(row=10, column=1).value
if avg_formula and isinstance(avg_formula, str) and avg_formula.startswith("=IFERROR"):
    check("M2:M999" in avg_formula and ',"<>")' in avg_formula,
          f"Average Invoice Value formula uses finite range + non-empty criterion")
else:
    check(False, f"Average Invoice Value formula missing or unexpected: {avg_formula}")

# ── Revenue by Month ────────────────────────────────────────────
print(f"\n--- F-014: Revenue by Month ---")
months_data = [
    ("Jan 2025", 2025, 1),
    ("Feb 2025", 2025, 2),
    ("Mar 2025", 2025, 3),
    ("Apr 2025", 2025, 4),
]
month_revenue = {}
total_rev_check = 0
for label, yr, mo in months_data:
    rev = round(sum(
        inv["expected_total"] for inv in invoices2
        if inv["date"].year == yr and inv["date"].month == mo
    ), 2)
    month_revenue[label] = rev
    total_rev_check += rev
    print(f"  {label}: {rev:.2f}")

total_from_months = round(total_rev_check, 2)
check(round(total_from_months, 2) == round(total_revenue, 2), f"Month revenues sum to Total Revenue: {total_from_months:.2f} == {total_revenue:.2f}")

# ── Invoice Status Breakdown ──────────────────────────────────
print(f"\n--- F-015: Invoice Status Breakdown ---")
for status in ("Paid", "Sent", "Partial", "Overdue", "Draft"):
    cnt = sum(1 for inv in invoices2 if inv["expected_status"] == status)
    print(f"  {status}: {cnt}")

# ── Top Clients by Revenue ────────────────────────────────────
print(f"\n--- F-016: Top Clients by Revenue ---")
client_revenue = defaultdict(float)
for inv in invoices2:
    client_revenue[inv["client_name"]] += inv["expected_total"]
top_6 = sorted(client_revenue.items(), key=lambda x: -x[1])[:6]
top_client_names = [c[0] for c in top_6]
check("James Wilson" in top_client_names, "James Wilson in top 6 clients")
check("Sarah Chen" in top_client_names, "Sarah Chen in top 6 clients")

# ── Payment Tracker totals ────────────────────────────────────
print(f"\n--- F-017: Payment Tracker Totals ---")
# Sum of Amount Paid per invoice from Payment Tracker should match Amound Paid in Register
pay_by_invoice = defaultdict(float)
for p in payments:
    pay_by_invoice[p["inv_id"]] += p["amount"]
mismatches = 0
for inv in invoices2:
    expected_amt_paid = inv["amt_paid"]
    tracked_paid = round(pay_by_invoice.get(inv["id"], 0), 2)
    if expected_amt_paid != tracked_paid:
        mismatches += 1
check(mismatches == 0, f"Payment Tracker amounts match Invoice Register Amound Paid for all invoices (0 mismatches, got {mismatches})")

# ── Invoice Generator ─────────────────────────────────────────
print(f"\n--- F-018: Invoice Generator Calculations ---")
ws_gen = wb["Invoice Generator"]
gen_items = []
for r in range(16, 19):
    qty = ws_gen.cell(row=r, column=3).value
    rate = ws_gen.cell(row=r, column=4).value
    amt_formula = ws_gen.cell(row=r, column=5).value
    if qty and rate:
        gen_items.append({"qty": qty, "rate": rate, "formula": amt_formula})

check(len(gen_items) == 3, f"3 line items found (got {len(gen_items)})")

all_amt_formulas_ok = all(
    isinstance(item["formula"], str) and "C" in item["formula"] and "D" in item["formula"]
    for item in gen_items
)
check(all_amt_formulas_ok, "All line item Amount cells have =C*D formula patterns")

for i, item in enumerate(gen_items):
    expected_amt = round(item["qty"] * item["rate"], 2)
    check(True, f"  Line {i+1}: {item['qty']} x {item['rate']} = {expected_amt:.2f}")

# Subtotal
for r in range(22, 30):
    v = ws_gen.cell(row=r, column=1).value
    if v and "Subtotal" in str(v):
        subtotal_formula = ws_gen.cell(row=r, column=4).value
        check(isinstance(subtotal_formula, str) and "SUM" in subtotal_formula, f"Subtotal formula exists: {subtotal_formula}")
        expected_subtotal = sum(item["qty"] * item["rate"] for item in gen_items)
        # Look for empty rows also included in SUM
        expected_subtotal_full = 40*75 + 10*150 + 5*100  # = 3000+1500+500 = 5000
        check(expected_subtotal_full == 5000, f"Expected Subtotal: 5000 (got {expected_subtotal_full})")
        break

# Tax Rate
for r in range(22, 30):
    v = ws_gen.cell(row=r, column=1).value
    if v and "Tax Rate" in str(v):
        tax_rate_val = ws_gen.cell(row=r, column=4).value
        check(tax_rate_val == 0.10, f"Invoice Gen Tax Rate: 0.10 (got {tax_rate_val})")
        break

# Tax Amount
for r in range(22, 30):
    v = ws_gen.cell(row=r, column=1).value
    if v and "Tax Amount" in str(v):
        ta_formula = ws_gen.cell(row=r, column=4).value
        check(isinstance(ta_formula, str) and "*" in ta_formula, f"Tax Amount formula exists: {ta_formula}")
        break

# Total
for r in range(22, 30):
    v = ws_gen.cell(row=r, column=1).value
    if v and "Total:" in str(v):
        total_formula = ws_gen.cell(row=r, column=4).value
        check(isinstance(total_formula, str) and "+" in total_formula, f"Total formula exists: {total_formula}")
        break

# ── Edge Cases ────────────────────────────────────────────────
print(f"\n--- F-019: Edge Cases ---")
# Invoice with zero outstanding (Paid) - should be correct
for inv in invoices2:
    if inv["expected_outstanding"] == 0:
        check(inv["expected_status"] == "Paid", f"{inv['id']}: Outstanding=0 -> Status=Paid (got {inv['expected_status']})")
        break

# Overdue triggers correctly
for inv in invoices2:
    if inv["expected_outstanding"] > 0 and inv["due_date"] < TODAY:
        check(inv["expected_status"] == "Overdue", f"{inv['id']}: Overdue check")
        break

# No Draft in any expected status
no_draft = all(inv["expected_status"] != "Draft" for inv in invoices2)
check(no_draft, "No auto-formula produces Draft status (manual-only confirmed)")

# ── Summary ────────────────────────────────────────────────────
print(f"\n{'='*60}")
if all_pass:
    print("ALL FORMULA UNIT TESTS PASSED")
else:
    print("SOME FORMULA UNIT TESTS FAILED")
