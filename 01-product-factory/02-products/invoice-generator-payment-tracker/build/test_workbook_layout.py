import openpyxl

path = r"C:\Users\sar1s\Documents\MyWork\Git\digital-lab\01-product-factory\02-products\invoice-generator-payment-tracker\build\nivora-invoice-generator-payment-tracker.xlsx"
wb = openpyxl.load_workbook(path, data_only=False)

all_pass = True
def check(ok, msg):
    global all_pass
    if ok:
        print(f"  PASS: {msg}")
    else:
        print(f"  FAIL: {msg}")
        all_pass = False

def get_text_in_range(ws, min_row, max_row, min_col, max_col):
    texts = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                texts.append(v)
    return texts

print("=" * 60)
print("  LAYOUT & UI QA")
print("=" * 60)

# ── Start Here ──────────────────────────────────────────────────
print(f"\n--- LAY-001: Start Here ---")
ws = wb["Start Here"]
texts = get_text_in_range(ws, 1, 50, 1, 3)
all_texts = " ".join(texts)

check("Nivora" in all_texts, "Product name 'Nivora' present")
check("Invoice Generator" in all_texts, "'Invoice Generator' in Start Here")
check("Payment Tracker" in all_texts, "'Payment Tracker' in Start Here")
check("Settings" in all_texts, "Settings tab mentioned")
check("Clients" in all_texts, "Clients tab mentioned")
check("Dashboard" in all_texts, "Dashboard tab mentioned")
check("Invoice Register" in all_texts, "Invoice Register tab mentioned")

# Check for setup instructions
check("Quick Start" in all_texts or any("1." in t for t in texts), "Setup/Quick Start instructions present")

# Check for important notes
check("formula" in all_texts.lower() and "do not edit" in all_texts.lower(), "Formula cell warning present")
check("backup" in all_texts.lower(), "Backup note present")
check("Google Sheets" in all_texts, "Google Sheets mention present")
check("Draft" in all_texts, "Draft status mentioned")

# Check no debug text
debug_phrases = ["TODO", "FIXME", "DEBUG", "REMOVE", "DELETE", "WIP", "XXX", "test data", "sample only"]
for phrase in debug_phrases:
    if phrase.lower() in all_texts.lower():
        check(False, f"No debug text '{phrase}' in Start Here")
        break
else:
    check(True, "No debug/internal text in Start Here")

# ── Dashboard ──────────────────────────────────────────────────
print(f"\n--- LAY-002: Dashboard ---")
ws = wb["Dashboard"]
texts = get_text_in_range(ws, 1, 50, 1, 8)
all_texts = " ".join(texts)

kpi_labels = ["Total Revenue", "Paid Revenue", "Outstanding Revenue", "Total Invoices",
              "Paid Invoices", "Unpaid Invoices", "Overdue Invoices", "Active Clients",
              "Average Invoice Value"]
for label in kpi_labels:
    check(label in all_texts, f"Dashboard KPI label: '{label}'")

section_labels = ["Revenue by Month", "Invoice Status Breakdown", "Top Clients by Revenue"]
for label in section_labels:
    check(label in all_texts, f"Dashboard section: '{label}'")

# ── Clients ─────────────────────────────────────────────────────
print(f"\n--- LAY-003: Clients ---")
ws = wb["Clients"]
expected_h = ["Client ID", "Client Name", "Company", "Email", "Phone", "Address", "Status", "Notes"]
for i, exp in enumerate(expected_h, 1):
    check(ws.cell(row=1, column=i).value == exp, f"Clients header col {i}: '{exp}'")
check(ws.freeze_panes == "A2", "Clients freeze panes at A2")

# ── Invoice Register ────────────────────────────────────────────
print(f"\n--- LAY-004: Invoice Register ---")
ws = wb["Invoice Register"]
expected_h = ["Invoice ID", "Invoice Date", "Due Date", "Client ID", "Client Name",
              "Description", "Subtotal", "Tax Rate", "Tax Amount", "Total",
              "Amount Paid", "Outstanding", "Status"]
for i, exp in enumerate(expected_h, 1):
    check(ws.cell(row=1, column=i).value == exp, f"Register header col {i}: '{exp}'")
check(ws.freeze_panes == "A2", "Register freeze panes at A2")

# ── Payment Tracker ─────────────────────────────────────────────
print(f"\n--- LAY-005: Payment Tracker ---")
ws = wb["Payment Tracker"]
expected_h = ["Payment ID", "Invoice ID", "Client Name", "Payment Date",
              "Payment Method", "Amount", "Notes"]
for i, exp in enumerate(expected_h, 1):
    check(ws.cell(row=1, column=i).value == exp, f"Payment header col {i}: '{exp}'")
check(ws.freeze_panes == "A2", "Payment freeze panes at A2")

# ── Invoice Generator ──────────────────────────────────────────
print(f"\n--- LAY-006: Invoice Generator ---")
ws = wb["Invoice Generator"]
texts = get_text_in_range(ws, 1, 50, 1, 5)
all_texts = " ".join(texts)

check("INVOICE" in all_texts, "Invoice title 'INVOICE' present")
check("From" in all_texts, "'From' section present")
check("Bill To" in all_texts, "'Bill To' section present")
check("Line Items" in all_texts, "'Line Items' section present")
check("Description" in all_texts, "'Description' column header present")
check("Quantity" in all_texts, "'Quantity' column header present")
check("Rate" in all_texts, "'Rate' column header present")
check("Summary" in all_texts, "'Summary' section present")
check("Subtotal" in all_texts, "'Subtotal' in Summary")
check("Tax Rate" in all_texts, "'Tax Rate' in Summary")
check("Tax Amount" in all_texts, "'Tax Amount' in Summary")
check("Total" in all_texts, "'Total' in Summary")
check("Payment" in all_texts, "'Payment Instructions' section present")
check("Invoice Number" in all_texts or all_texts.count("Invoice") >= 2, "Invoice Number field present")
check("Invoice Date" in all_texts or "Date:" in all_texts, "Invoice Date field present")
check("Due Date" in all_texts or "Due" in all_texts, "Due Date field present")
check("Your Name" in all_texts or "Company" in all_texts, "'Your Name/Company' field present")
check("Client Name" in all_texts, "'Client Name' field present")
check("Address" in all_texts, "'Address' field present")

# Check no debug text
debug_phrases = ["TODO", "FIXME", "DEBUG", "REMOVE", "DELETE", "WIP", "XXX"]
for phrase in debug_phrases:
    if phrase.lower() in all_texts.lower():
        check(False, f"No debug text '{phrase}' in Invoice Generator")
        break
else:
    check(True, "No debug/internal text in Invoice Generator")

# ── Settings ────────────────────────────────────────────────────
print(f"\n--- LAY-007: Settings ---")
ws = wb["Settings"]
texts = get_text_in_range(ws, 1, 40, 1, 2)
all_texts = " ".join(texts)

check("Currency Symbol" in all_texts, "Currency Symbol setting present")
check("Default Tax Rate" in all_texts, "Default Tax Rate setting present")
check("Invoice Status Values" in all_texts, "Invoice Status Values section present")
check("Payment Method Values" in all_texts, "Payment Method Values section present")
check("Client Status Values" in all_texts, "Client Status Values section present")

# Check Draft is in Status Values
draft_found = any(
    ws.cell(row=r, column=1).value == "Draft"
    for r in range(1, 40)
)
check(draft_found, "Draft listed in Settings Status Values")

# ── Frozen panes ────────────────────────────────────────────────
print(f"\n--- LAY-008: Freeze panes ---")
check(wb["Clients"].freeze_panes == "A2", "Clients: freeze")
check(wb["Invoice Register"].freeze_panes == "A2", "Register: freeze")
check(wb["Payment Tracker"].freeze_panes == "A2", "Payment: freeze")
check(wb["Start Here"].freeze_panes is None, "Start Here: no freeze")
check(wb["Dashboard"].freeze_panes is None, "Dashboard: no freeze")
check(wb["Invoice Generator"].freeze_panes is None, "Invoice Gen: no freeze")

# ── Number formats on key cells ─────────────────────────────────
print(f"\n--- LAY-009: Number formatting ---")
ws_reg = wb["Invoice Register"]
date_fmts = set()
curr_fmts = set()
for r in range(2, 27):
    date_fmts.add(ws_reg.cell(row=r, column=2).number_format)
    date_fmts.add(ws_reg.cell(row=r, column=3).number_format)
    curr_fmts.add(ws_reg.cell(row=r, column=7).number_format)
    curr_fmts.add(ws_reg.cell(row=r, column=9).number_format)

has_date_format = any("MMM" in str(f).upper() or "DATE" in str(f).upper() for f in date_fmts)
has_currency_format = any(("#,##0" in str(f) or "0.00" in str(f)) for f in curr_fmts)
check(has_date_format, f"Date formatting on date columns: {date_fmts}")
check(has_currency_format, f"Currency formatting on money columns: {curr_fmts}")

# ── Header row styling ──────────────────────────────────────────
print(f"\n--- LAY-010: Header styling ---")
for sname in ["Clients", "Invoice Register", "Payment Tracker"]:
    ws = wb[sname]
    h_cell = ws.cell(row=1, column=1)
    is_bold = h_cell.font and h_cell.font.bold
    check(is_bold, f"{sname}: Header row is bold")

# ── Summary ────────────────────────────────────────────────────
print(f"\n{'='*60}")
if all_pass:
    print("ALL LAYOUT QA CHECKS PASSED")
else:
    print("SOME LAYOUT QA CHECKS FAILED")
